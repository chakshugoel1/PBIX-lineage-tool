"""
excel_report.py

Generate the Model Change Impact Excel workbook from a snapshot pair, a
diff result (diff.py), an impact result (impact.py), and the changed
file's report layout (report_layout.py). Nothing here is specific to any
one model - every sheet is built purely from the generic dicts those other
modules produce, so the same code works for any baseline/changed PBIX pair.

Sheets: Summary, Changed Tables, Changed Measures, Changed Columns, Changed
Relationships, Visual Impact, KPI Impact, Playwright Input, Manual Review,
Object Inventory.

`build_excel_report()` is the only entry point most callers need.
"""
import difflib
import os
import re

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font, PatternFill, Alignment

from services import fileutils

HEADER_FILL = PatternFill(start_color="FF83CCEB", end_color="FF83CCEB", fill_type="solid")
RED_FILL = PatternFill(start_color="FFFF4B4B", end_color="FFFF4B4B", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
GREY_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

# Word-level diff highlighting for Before/After cells (see _diff_cell_values).
_TOKEN_RE = re.compile(r"\s+|\w+|[^\w\s]")
_MAX_DIFF_TOKENS = 400
_REMOVED_FONT = InlineFont(color="FFCC0000", strike=True)
_ADDED_FONT = InlineFont(color="FF107C10", b=True)

_TABLE_FIELDS = ("is_calculated_table", "m_expression", "is_hidden", "description")
_MEASURE_FIELDS = ("expression", "display_folder", "description", "format_string", "is_hidden", "kpi_id")
_COLUMN_FIELDS = ("data_type", "is_calculated", "expression", "format_string", "is_hidden", "description", "display_folder")
_RELATIONSHIP_FIELDS = ("is_active", "cardinality", "cross_filtering_behavior", "rely_on_referential_integrity")


def build_excel_report(baseline_snapshot, changed_snapshot, diff_result, impact_result, report_layout, output_path):
    """Build the full Model Change Impact workbook and write it to
    `output_path` (atomically, via the same helper V1 uses). Returns
    `output_path`."""
    visual_rows = _flatten_visual_impacts(impact_result)
    manual_review_rows = _build_manual_review_rows(diff_result, report_layout, visual_rows)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, diff_result, report_layout, visual_rows, manual_review_rows)

    _write_table_sheet(wb.create_sheet("Changed Tables"), diff_result["tables"])
    _write_measure_sheet(wb.create_sheet("Changed Measures"), diff_result["measures"])
    _write_column_sheet(wb.create_sheet("Changed Columns"), diff_result["columns"])
    _write_relationship_sheet(wb.create_sheet("Changed Relationships"), diff_result["relationships"])
    _write_visual_impact_sheet(wb.create_sheet("Visual Impact"), visual_rows)
    _write_kpi_impact_sheet(wb.create_sheet("KPI Impact"), visual_rows)
    _write_playwright_input_sheet(wb.create_sheet("Playwright Input"), visual_rows)
    _write_manual_review_sheet(wb.create_sheet("Manual Review"), manual_review_rows)
    _write_object_inventory_sheet(wb.create_sheet("Object Inventory"), changed_snapshot)

    output_dir = os.path.dirname(os.path.abspath(output_path))
    if output_dir and not os.path.isdir(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    fileutils.atomic_replace_workbook(wb, output_path)
    return output_path


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _fmt(value, max_len=32000):
    """Stringify a field value for display. `max_len` is a hard safety cap at
    Excel's per-cell character limit (32,767), not a normal truncation -
    real DAX/M expressions are always shown in full."""
    if value is None:
        return "(none)"
    text = str(value)
    return text if len(text) <= max_len else text[:max_len] + "... [truncated - exceeds Excel's cell limit]"


def _full_fields_text(item, fields):
    return "\n".join(f"{f}: {_fmt(item.get(f))}" for f in fields)


def _diff_cell_values(before, after):
    """Return (before_value, after_value) ready to assign to openpyxl cells:
    full, untruncated text for both sides, with the changed tokens
    highlighted (red strikethrough on the before side, bold green on the
    after side) via rich text. Falls back to plain text if there's nothing
    to diff or the text is too large to diff usefully."""
    before_text, after_text = _fmt(before), _fmt(after)
    if before_text == after_text:
        return before_text, after_text
    before_tokens, after_tokens = _TOKEN_RE.findall(before_text), _TOKEN_RE.findall(after_text)
    if len(before_tokens) > _MAX_DIFF_TOKENS or len(after_tokens) > _MAX_DIFF_TOKENS:
        return before_text, after_text
    matcher = difflib.SequenceMatcher(a=before_tokens, b=after_tokens, autojunk=False)
    before_parts, after_parts = [], []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        b_chunk, a_chunk = "".join(before_tokens[i1:i2]), "".join(after_tokens[j1:j2])
        if tag == "equal":
            before_parts.append(b_chunk)
            after_parts.append(a_chunk)
        else:
            if b_chunk:
                before_parts.append(TextBlock(_REMOVED_FONT, b_chunk))
            if a_chunk:
                after_parts.append(TextBlock(_ADDED_FONT, a_chunk))
    before_value = CellRichText(before_parts) if any(isinstance(p, TextBlock) for p in before_parts) else before_text
    after_value = CellRichText(after_parts) if any(isinstance(p, TextBlock) for p in after_parts) else after_text
    return before_value, after_value


def _write_header(ws, headers):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _apply_wrap(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP


def _cell_text(value):
    """Plain-text rendering of a cell value, including CellRichText (whose
    default str()/list repr is not the text Excel actually displays)."""
    if isinstance(value, CellRichText):
        return "".join(str(part) for part in value)
    return str(value)


def _autofit_columns(ws, min_width=10, max_width=60):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            longest = max((len(line) for line in _cell_text(cell.value).splitlines()), default=0)
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), longest)
    for col_letter, longest in widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, longest + 2))


def _describe_changed_object(kind, detail):
    if kind == "relationship":
        ident = detail.get("identity_after", detail)
        return f"{ident.get('from_table')}[{ident.get('from_column')}] -> {ident.get('to_table')}[{ident.get('to_column')}]"
    ident = detail.get("identity_after", detail)
    if kind == "table":
        return ident.get("table", "")
    return f"{ident.get('table', '')}[{ident.get('name', '')}]"


def _impact_basis(kind, change_type, matched_via):
    """Classify how confident this visual<->change link is. Table/
    relationship changes deliberately over-include (every column of the
    affected table is treated as a possible dependent - see impact.py's
    `_table_column_keys` seeding) so this label lets the report - and the
    reader - tell a precise hit from a broad, heuristic one."""
    if kind in ("measure", "column"):
        return "Direct" if matched_via == "direct" else "Dependency chain"
    if change_type == "removed" and matched_via == "direct":
        return "Direct (references removed object)"
    return "Broad (table/relationship-level)"


_BASIS_RANK = {
    "Direct": 1,
    "Direct (references removed object)": 1,
    "Dependency chain": 2,
    "Broad (table/relationship-level)": 3,
}


def _flatten_visual_impacts(impact_result):
    """One row per (page, visual), aggregating every changed object that
    touches it - a visual can be pulled in by more than one root-cause
    change, and previously got one row per cause, which is what made this
    sheet balloon. Sorted by page so impact is easy to scan page-by-page."""
    by_visual = {}
    for section, kind in (
        ("measures", "measure"), ("columns", "column"), ("tables", "table"), ("relationships", "relationship"),
    ):
        for record in impact_result[section]:
            changed_object = _describe_changed_object(kind, record["detail"])
            for visual in record["impacted_visuals"]:
                basis = _impact_basis(kind, record["change_type"], visual["matched_via"])
                key = (visual["page_id"], visual["visual_id"])
                row = by_visual.setdefault(key, {
                    "page_id": visual["page_id"],
                    "page_display_name": visual["page_display_name"],
                    "visual_id": visual["visual_id"],
                    "visual_type": visual["visual_type"],
                    "kpi_classification": visual["kpi_classification"],
                    "reasons": [],
                })
                row["reasons"].append({
                    "kind": kind,
                    "change_type": record["change_type"],
                    "changed_object": changed_object,
                    "matched_via": visual["matched_via"],
                    "matched_object": visual["matched_object"],
                    "impact_basis": basis,
                })

    rows = list(by_visual.values())
    for row in rows:
        row["reasons"].sort(key=lambda r: _BASIS_RANK[r["impact_basis"]])
        row["best_basis"] = row["reasons"][0]["impact_basis"]
    rows.sort(key=lambda r: (r["page_display_name"] or r["page_id"] or "", r["visual_id"] or ""))
    return rows


def _reasons_text(reasons):
    return "\n".join(
        f"[{r['impact_basis']}] {r['kind']} {r['change_type']}: {r['changed_object']}" for r in reasons
    )


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def _write_summary_sheet(ws, diff_result, report_layout, visual_rows, manual_review_rows):
    ws.cell(row=1, column=1, value="Model Change Impact Report").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Baseline file:")
    ws.cell(row=2, column=2, value=diff_result.get("baseline_file"))
    ws.cell(row=3, column=1, value="Changed file:")
    ws.cell(row=3, column=2, value=diff_result.get("changed_file"))
    ws.cell(row=4, column=1, value="Report layout format:")
    ws.cell(row=4, column=2, value=report_layout.get("format"))

    row = 6
    for c, h in enumerate(("Entity", "Added", "Removed", "Changed", "Rename Candidates", "Unchanged", "Detail Sheet"), start=1):
        ws.cell(row=row, column=c, value=h).font = Font(bold=True)
    row += 1

    for section, sheet_name in (
        ("tables", "Changed Tables"), ("measures", "Changed Measures"),
        ("columns", "Changed Columns"), ("relationships", "Changed Relationships"),
    ):
        s = diff_result[section]
        rename_count = sum(1 for c in s["changed"] if c["is_rename_candidate"])
        ws.cell(row=row, column=1, value=section.capitalize())
        ws.cell(row=row, column=2, value=len(s["added"]))
        ws.cell(row=row, column=3, value=len(s["removed"]))
        ws.cell(row=row, column=4, value=len(s["changed"]))
        ws.cell(row=row, column=5, value=rename_count)
        ws.cell(row=row, column=6, value=s["unchanged_count"])
        _add_sheet_link(ws, row, 7, sheet_name)
        row += 1

    unique_visuals = {(r["page_id"], r["visual_id"]) for r in visual_rows}
    unique_kpi_visuals = {(r["page_id"], r["visual_id"]) for r in visual_rows if r["kpi_classification"] in ("certain", "heuristic")}
    broad_only_count = sum(1 for r in visual_rows if r["best_basis"] == "Broad (table/relationship-level)")

    row += 1
    ws.cell(row=row, column=1, value="Impacted visuals (unique):")
    ws.cell(row=row, column=2, value=len(unique_visuals))
    _add_sheet_link(ws, row, 7, "Visual Impact")
    row += 1
    ws.cell(row=row, column=1, value="  - direct / confirmed:")
    ws.cell(row=row, column=2, value=len(unique_visuals) - broad_only_count)
    row += 1
    ws.cell(row=row, column=1, value="  - broad / heuristic-level only (review before acting on these):")
    ws.cell(row=row, column=2, value=broad_only_count)
    row += 1
    ws.cell(row=row, column=1, value="Impacted KPI/card visuals (unique):")
    ws.cell(row=row, column=2, value=len(unique_kpi_visuals))
    _add_sheet_link(ws, row, 7, "KPI Impact")
    row += 1
    ws.cell(row=row, column=1, value="Playwright test candidates:")
    ws.cell(row=row, column=2, value=len(visual_rows))
    _add_sheet_link(ws, row, 7, "Playwright Input")
    row += 1
    ws.cell(row=row, column=1, value="Manual review items:")
    ws.cell(row=row, column=2, value=len(manual_review_rows))
    _add_sheet_link(ws, row, 7, "Manual Review")

    _autofit_columns(ws)


def _add_sheet_link(ws, row, col, sheet_name):
    cell = ws.cell(row=row, column=col, value=f"Go to {sheet_name}")
    cell.hyperlink = f"#'{sheet_name}'!A1"
    cell.style = "Hyperlink"


# ---------------------------------------------------------------------------
# Changed Tables / Measures / Columns / Relationships
# ---------------------------------------------------------------------------

def _write_table_sheet(ws, tables_diff):
    _write_header(ws, ["Table", "Change Type", "Rename Candidate", "Field", "Before", "After"])
    for item in tables_diff["added"]:
        ws.append([item["table"], "Added", "", "(all fields)", "(did not exist)", _full_fields_text(item, _TABLE_FIELDS)])
    for item in tables_diff["removed"]:
        ws.append([item["table"], "Removed", "", "(all fields)", _full_fields_text(item, _TABLE_FIELDS), "(no longer exists)"])
    for item in tables_diff["changed"]:
        name = item["identity_after"]["table"]
        if item["is_rename_candidate"]:
            name += f" (was: {item['identity_before']['table']})"
        change_type = "Renamed" if item["is_rename_candidate"] else "Modified"
        rename_flag = "Yes" if item["is_rename_candidate"] else ""
        changed_fields = [f for f in _TABLE_FIELDS if f in item["field_changes"]]
        if not changed_fields:
            ws.append([name, change_type, rename_flag, "(name only)",
                       item["identity_before"]["table"], item["identity_after"]["table"]])
            continue
        for field in changed_fields:
            before_val, after_val = _diff_cell_values(
                item["field_changes"][field]["before"], item["field_changes"][field]["after"])
            ws.append([name, change_type, rename_flag, field, before_val, after_val])
    _apply_wrap(ws)
    _autofit_columns(ws)


def _write_measure_sheet(ws, measures_diff):
    _write_header(ws, ["Table", "Measure", "Change Type", "Rename Candidate", "Field", "Before", "After"])
    for item in measures_diff["added"]:
        ws.append([item["table"], item["name"], "Added", "", "(all fields)", "(did not exist)",
                   _full_fields_text(item, _MEASURE_FIELDS)])
    for item in measures_diff["removed"]:
        ws.append([item["table"], item["name"], "Removed", "", "(all fields)",
                   _full_fields_text(item, _MEASURE_FIELDS), "(no longer exists)"])
    for item in measures_diff["changed"]:
        name = item["identity_after"]["name"]
        if item["is_rename_candidate"]:
            name += f" (was: {item['identity_before']['name']})"
        change_type = "Renamed" if item["is_rename_candidate"] else "Modified"
        rename_flag = "Yes" if item["is_rename_candidate"] else ""
        table = item["identity_after"]["table"]
        changed_fields = [f for f in _MEASURE_FIELDS if f in item["field_changes"]]
        if not changed_fields:
            ws.append([table, name, change_type, rename_flag, "(name only)",
                       item["identity_before"]["name"], item["identity_after"]["name"]])
            continue
        for field in changed_fields:
            before_val, after_val = _diff_cell_values(
                item["field_changes"][field]["before"], item["field_changes"][field]["after"])
            ws.append([table, name, change_type, rename_flag, field, before_val, after_val])
    _apply_wrap(ws)
    _autofit_columns(ws)


def _write_column_sheet(ws, columns_diff):
    _write_header(ws, ["Table", "Column", "Change Type", "Rename Candidate", "Field", "Before", "After"])
    for item in columns_diff["added"]:
        ws.append([item["table"], item["name"], "Added", "", "(all fields)", "(did not exist)",
                   _full_fields_text(item, _COLUMN_FIELDS)])
    for item in columns_diff["removed"]:
        ws.append([item["table"], item["name"], "Removed", "", "(all fields)",
                   _full_fields_text(item, _COLUMN_FIELDS), "(no longer exists)"])
    for item in columns_diff["changed"]:
        before, after = item["identity_before"], item["identity_after"]
        name = after["name"]
        if item["is_rename_candidate"]:
            name += f" (was: {before['table']}[{before['name']}])"
        change_type = "Renamed" if item["is_rename_candidate"] else "Modified"
        rename_flag = "Yes" if item["is_rename_candidate"] else ""
        changed_fields = [f for f in _COLUMN_FIELDS if f in item["field_changes"]]
        if not changed_fields:
            ws.append([after["table"], name, change_type, rename_flag, "(name only)",
                       f"{before['table']}[{before['name']}]", f"{after['table']}[{after['name']}]"])
            continue
        for field in changed_fields:
            before_val, after_val = _diff_cell_values(
                item["field_changes"][field]["before"], item["field_changes"][field]["after"])
            ws.append([after["table"], name, change_type, rename_flag, field, before_val, after_val])
    _apply_wrap(ws)
    _autofit_columns(ws)


def _write_relationship_sheet(ws, rel_diff):
    _write_header(ws, ["From Table", "From Column", "To Table", "To Column", "Change Type", "Field", "Before", "After"])
    for item in rel_diff["added"]:
        ws.append([item["from_table"], item["from_column"], item["to_table"], item["to_column"], "Added",
                   "(all fields)", "(did not exist)", _full_fields_text(item, _RELATIONSHIP_FIELDS)])
    for item in rel_diff["removed"]:
        ws.append([item["from_table"], item["from_column"], item["to_table"], item["to_column"], "Removed",
                   "(all fields)", _full_fields_text(item, _RELATIONSHIP_FIELDS), "(no longer exists)"])
    for item in rel_diff["changed"]:
        ident = item["identity_after"]
        for field in (f for f in _RELATIONSHIP_FIELDS if f in item["field_changes"]):
            before_val, after_val = _diff_cell_values(
                item["field_changes"][field]["before"], item["field_changes"][field]["after"])
            ws.append([ident["from_table"], ident["from_column"], ident["to_table"], ident["to_column"],
                       "Modified", field, before_val, after_val])
    _apply_wrap(ws)
    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Visual Impact / KPI Impact / Playwright Input
# ---------------------------------------------------------------------------

def _write_visual_impact_sheet(ws, visual_rows):
    _write_header(ws, ["Page", "Visual ID", "Visual Type", "KPI Classification", "Impact Basis", "Reasons (changed object(s))"])
    ws.auto_filter.ref = "A1:F1"
    current_page = object()
    for r in visual_rows:
        page_label = r["page_display_name"] or r["page_id"]
        if page_label != current_page:
            current_page = page_label
            header_row = ws.max_row + 1
            ws.cell(row=header_row, column=1, value=f"Page: {page_label}").font = Font(bold=True)
            ws.cell(row=header_row, column=1).fill = HEADER_FILL
            ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=6)
        ws.append([page_label, r["visual_id"], r["visual_type"], r["kpi_classification"] or "",
                   r["best_basis"], _reasons_text(r["reasons"])])
        if r["best_basis"] == "Broad (table/relationship-level)":
            for cell in ws[ws.max_row]:
                cell.fill = GREY_FILL
    _apply_wrap(ws)
    _autofit_columns(ws)


def _write_kpi_impact_sheet(ws, visual_rows):
    kpi_rows = [r for r in visual_rows if r["kpi_classification"] in ("certain", "heuristic")]
    _write_header(ws, ["Page", "Visual ID", "Visual Type", "KPI Confidence", "Impact Basis", "Reasons (changed object(s))"])
    ws.auto_filter.ref = "A1:F1"
    current_page = object()
    for r in kpi_rows:
        page_label = r["page_display_name"] or r["page_id"]
        if page_label != current_page:
            current_page = page_label
            header_row = ws.max_row + 1
            ws.cell(row=header_row, column=1, value=f"Page: {page_label}").font = Font(bold=True)
            ws.cell(row=header_row, column=1).fill = HEADER_FILL
            ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=6)
        ws.append([page_label, r["visual_id"], r["visual_type"], r["kpi_classification"],
                   r["best_basis"], _reasons_text(r["reasons"])])
        if r["kpi_classification"] == "heuristic":
            for cell in ws[ws.max_row]:
                cell.fill = YELLOW_FILL
        elif r["best_basis"] == "Broad (table/relationship-level)":
            for cell in ws[ws.max_row]:
                cell.fill = GREY_FILL
    _apply_wrap(ws)
    _autofit_columns(ws)


def _severity(row):
    if row["kpi_classification"] == "certain":
        return "High"
    if row["best_basis"] in ("Direct", "Direct (references removed object)"):
        return "Medium"
    return "Low"


def _recommended_assertion(row):
    changed_summary = "; ".join(dict.fromkeys(reason["changed_object"] for reason in row["reasons"]))
    if row["kpi_classification"] in ("certain", "heuristic"):
        return (f"Verify KPI/card visual '{row['visual_id']}' on page '{row['page_display_name']}' still shows "
                f"the expected value/format after change(s) to: {changed_summary}.")
    return (f"Verify visual '{row['visual_id']}' on page '{row['page_display_name']}' still renders correctly "
            f"and its data is accurate after change(s) to: {changed_summary}.")


def _write_playwright_input_sheet(ws, visual_rows):
    _write_header(ws, ["Page", "Visual", "Visual Type", "Severity", "Impact Basis", "Changed Object(s)",
                        "Recommended Assertion", "Page ID / Visual ID"])
    ws.auto_filter.ref = "A1:H1"
    fill_by_severity = {"High": RED_FILL, "Medium": YELLOW_FILL, "Low": None}
    for r in visual_rows:
        severity = _severity(r)
        changed_objects = "\n".join(dict.fromkeys(reason["changed_object"] for reason in r["reasons"]))
        ws.append([
            r["page_display_name"] or r["page_id"], r["visual_id"], r["visual_type"], severity,
            r["best_basis"], changed_objects, _recommended_assertion(r),
            f"page={r['page_id']} / visual={r['visual_id']}",
        ])
        fill = fill_by_severity[severity]
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    _apply_wrap(ws)
    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Manual Review
# ---------------------------------------------------------------------------

def _build_manual_review_rows(diff_result, report_layout, visual_rows):
    rows = []
    if report_layout.get("format") != "pbir":
        rows.append(("Report Layout", report_layout.get("format"),
                     report_layout.get("unsupported_reason") or "Non-PBIR report layout format - visual/page detection is best-effort."))

    rel = diff_result["relationships"]
    for item in rel["added"] + rel["removed"]:
        rows.append(("Relationship", _describe_changed_object("relationship", item),
                     "Relationship changes have broad, heuristic impact - review affected visuals manually."))
    for item in rel["changed"]:
        reason = "Relationship changes have broad, heuristic impact - review affected visuals manually."
        if set(item["field_changes"]) == {"cardinality"}:
            reason = ("Only 'cardinality' changed, which Power BI often infers from the data itself rather "
                      "than an explicit relationship edit - this may just be a side effect of a data refresh. "
                      "Verify before treating it as an intentional design change.")
        rows.append(("Relationship", _describe_changed_object("relationship", item), reason))

    for item in diff_result["tables"]["changed"]:
        if "m_expression" in item["field_changes"]:
            rows.append(("Table (Power Query)", item["identity_after"]["table"],
                         "Power Query change - downstream column/measure impact was estimated heuristically "
                         "(every column of this table was treated as potentially affected)."))

    for section, label in (("tables", "Table"), ("measures", "Measure"), ("columns", "Column")):
        for item in diff_result[section]["changed"]:
            if item["is_rename_candidate"]:
                rows.append((f"{label} rename", f"{item['identity_before']} -> {item['identity_after']}",
                             "Matched via lineage_tag - confirm this is genuinely a rename, not a coincidence."))

    for r in visual_rows:
        if r["kpi_classification"] == "heuristic":
            rows.append(("KPI classification", r["visual_id"],
                         "Custom visual type name suggests a KPI/card, but this is a heuristic - confirm manually."))
        if any(reason["change_type"] == "removed" and reason["impact_basis"].startswith("Direct")
               for reason in r["reasons"]):
            removed_objects = ", ".join(dict.fromkeys(
                reason["changed_object"] for reason in r["reasons"] if reason["change_type"] == "removed"
            ))
            rows.append(("Dangling binding", r["visual_id"],
                         f"Visual still references a removed object ({removed_objects}) - likely broken, "
                         "verify the report was updated."))

    return rows


def _write_manual_review_sheet(ws, manual_review_rows):
    _write_header(ws, ["Category", "Item", "Reason"])
    for category, item, reason in manual_review_rows:
        ws.append([category, item, reason])
    _apply_wrap(ws)
    _autofit_columns(ws)


# ---------------------------------------------------------------------------
# Object Inventory
# ---------------------------------------------------------------------------

def _write_object_inventory_sheet(ws, changed_snapshot):
    _write_header(ws, ["Object Type", "Table", "Name", "Details"])
    for table_name, table in changed_snapshot.get("tables", {}).items():
        ws.append(["Table", table_name, "", f"is_calculated_table={table.get('is_calculated_table')}; is_hidden={table.get('is_hidden')}"])
        for col in table.get("columns", []):
            ws.append(["Column", table_name, col["name"],
                       f"data_type={col.get('data_type')}; is_calculated={col.get('is_calculated')}; is_hidden={col.get('is_hidden')}"])
    for table_name, measures in changed_snapshot.get("measures", {}).items():
        for name, measure in measures.items():
            ws.append(["Measure", table_name, name,
                       f"is_hidden={measure.get('is_hidden')}; format_string={measure.get('format_string')}"])
    for rel in changed_snapshot.get("relationships", []):
        ws.append(["Relationship", rel["from_table"], f"{rel['from_column']} -> {rel['to_table']}.{rel['to_column']}",
                   f"cardinality={rel.get('cardinality')}; is_active={rel.get('is_active')}"])
    _apply_wrap(ws)
    _autofit_columns(ws)
