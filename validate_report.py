import re
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
import config

TARGET = config.TARGET_XLSX
GENERATED = config.GENERATED_XLSX

# Handles: "Entity Name:", "Entity Name :" (extra space), "Entitty Name:" (typo),
# "Entity name:" (lowercase) - all found literally present in the source sheet.
RE_ENTITY_NAME = re.compile(r'(?:Entity|Entitty)\s*Name\s*:\s*(.+)', re.IGNORECASE)


def load_sheet_rows(path, sheet="CashPlus"):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = {}
    for r in range(2, ws.max_row + 1):  # row 1 is the column-header row
        c3 = ws.cell(row=r, column=3).value
        if not c3:
            continue
        m = RE_ENTITY_NAME.search(str(c3))
        if not m:
            name = str(c3).split("\n")[0].strip()
        else:
            name = m.group(1).strip()
        rows[name] = {
            "row": r,
            "entities_used": c3,
            "level1": ws.cell(row=r, column=4).value,
            "level2": ws.cell(row=r, column=5).value,
            "final_source": ws.cell(row=r, column=6).value,
            "folder": ws.cell(row=r, column=7).value,
            "file": ws.cell(row=r, column=8).value,
            "remarks": ws.cell(row=r, column=9).value,
        }
    return rows


def norm(s):
    if s is None:
        return ""
    return re.sub(r'\s+', ' ', str(s)).strip().lower()


def extract_dataflow(level_text):
    if not level_text:
        return None
    m = re.search(r'Dataflow:\s*([^\n]+)', level_text)
    return m.group(1).strip() if m else None


def extract_key_values(final_source):
    """Pull out datamart/schema/table/file tokens regardless of exact label formatting."""
    if not final_source:
        return {}
    text = str(final_source)
    out = {}
    m = re.search(r'Datamart\s*=\s*"?([^\n",]+)"?', text, re.IGNORECASE)
    if m:
        out["datamart"] = m.group(1).strip()
    m = re.search(r'Schema\s*=\s*"?([^\n",]+)"?', text, re.IGNORECASE)
    if m:
        out["schema"] = m.group(1).strip()
    m = re.search(r'Item\s*=\s*"?([^\n",]+)"?', text, re.IGNORECASE)
    if m:
        out["item"] = m.group(1).strip()
    m = re.search(r'(?:XLSX Name|File Name)\s*=\s*"?([^\n",]+)"?', text, re.IGNORECASE)
    if m:
        out["file"] = m.group(1).strip()
    return out


target = load_sheet_rows(TARGET)
generated = load_sheet_rows(GENERATED)

print(f"Target rows: {len(target)}   Generated rows: {len(generated)}")

only_target = set(target) - set(generated)
only_generated = set(generated) - set(target)
common = set(target) & set(generated)

print(f"Only in target: {len(only_target)}  Only in generated: {len(only_generated)}  Common: {len(common)}")

match_l1 = match_l2 = match_final = 0
mismatches = []

for name in sorted(common):
    t = target[name]
    g = generated[name]
    t_df1 = extract_dataflow(t["level1"])
    g_df1 = extract_dataflow(g["level1"])
    l1_ok = norm(t_df1) == norm(g_df1) or (not t_df1 and not g_df1)
    if l1_ok:
        match_l1 += 1

    t_df2 = extract_dataflow(t["level2"])
    g_df2 = extract_dataflow(g["level2"])
    l2_ok = norm(t_df2) == norm(g_df2) or (not t_df2 and not g_df2)
    if l2_ok:
        match_l2 += 1

    t_kv = extract_key_values(t["final_source"])
    g_kv = extract_key_values(g["final_source"])
    final_ok = True
    for k in ("datamart", "schema", "item", "file"):
        if t_kv.get(k) and norm(t_kv.get(k)) != norm(g_kv.get(k)):
            final_ok = False
    if final_ok:
        match_final += 1

    if not (l1_ok and l2_ok and final_ok):
        mismatches.append((name, l1_ok, l2_ok, final_ok, t_df1, g_df1, t_df2, g_df2, t_kv, g_kv))

n = len(common)
print(f"\nLevel1 dataflow match: {match_l1}/{n} ({100*match_l1/n:.1f}%)")
print(f"Level2 dataflow match: {match_l2}/{n} ({100*match_l2/n:.1f}%)")
print(f"Final source key-values match: {match_final}/{n} ({100*match_final/n:.1f}%)")

print(f"\n--- Only in target ({len(only_target)}) ---")
for name in sorted(only_target):
    print(" ", name)

print(f"\n--- Only in generated ({len(only_generated)}) ---")
for name in sorted(only_generated):
    print(" ", name)

print(f"\n--- Mismatches ({len(mismatches)}) ---")
for name, l1_ok, l2_ok, final_ok, t_df1, g_df1, t_df2, g_df2, t_kv, g_kv in mismatches[:60]:
    print(f"\n[{name}]  L1_ok={l1_ok} L2_ok={l2_ok} FINAL_ok={final_ok}")
    if not l1_ok:
        print(f"    L1 target={t_df1!r}  generated={g_df1!r}")
    if not l2_ok:
        print(f"    L2 target={t_df2!r}  generated={g_df2!r}")
    if not final_ok:
        print(f"    FINAL target={t_kv}  generated={g_kv}")
