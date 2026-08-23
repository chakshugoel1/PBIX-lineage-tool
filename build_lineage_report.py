"""
build_lineage_report.py

Generates the final PBIX -> Dataflow -> Physical Source lineage workbook in
the same 3-sheet layout as CASHPLUS-DASHBOARD 1.xlsx ("CashPlus",
"Dataflows used", "Unused tables"), using pbixray + the dataflow JSON files
configured in config.py.

Setup (see README.md for full instructions):
  python -m venv .venv
  .venv/Scripts/Activate.ps1
  pip install -r requirements.txt
  python build_lineage_report.py
"""
import sys
import os
import re

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from pbixray import PBIXRay
import lineage_lib as ll
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from collections import Counter
import config
import build_transformations_report as btr

PBIX_PATH = config.PBIX_PATH
DATAFLOW_FOLDER = config.DATAFLOW_FOLDER
OUTPUT_PATH = config.GENERATED_XLSX

HEADER_FILL = PatternFill(start_color="FF83CCEB", end_color="FF83CCEB", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FFC6E0B4", end_color="FFC6E0B4", fill_type="solid")
RED_FILL = PatternFill(start_color="FFFF4B4B", end_color="FFFF4B4B", fill_type="solid")
GREY_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

HEADERS = ["Color Codes", "Report Name", "Entities Used across report", "Source  - Level1",
           "Source - Level 2", "Final Source", "Folder PATH", "File Name", "Remarks"]


def _safe_column_set(df, col):
    # pbixray returns a columnless DataFrame (not just 0 rows) when a table (e.g. measures) is absent from the PBIX.
    return set(df[col]) if col in df.columns else set()


def load_everything(pbix_path=None, dataflow_folder=None):
    pbix_path = pbix_path or PBIX_PATH
    dataflow_folder = dataflow_folder or DATAFLOW_FOLDER

    # Validate PBIX file exists before attempting to load
    if not os.path.isfile(pbix_path):
        raise FileNotFoundError(
            f"PBIX file not found: '{pbix_path}'\n"
            f"Check config.PBIX_PATH and ensure the file exists and is accessible."
        )

    model = PBIXRay(pbix_path)
    entries = {}
    for _, row in model.power_query.iterrows():
        entries[row["TableName"]] = str(row["Expression"])
    for _, row in model.m_parameters.iterrows():
        entries[row["ParameterName"]] = str(row["Expression"])

    pbix_universe = ll.Universe(entries)
    global_params = pbix_universe.build_global_param_values()
    guid_cache = ll.load_guid_cache()
    direct, enumerators, unrecognized = ll.analyze_direct_dataflow_bindings(pbix_universe, global_params, guid_cache)
    entity_of = ll.build_entity_of(pbix_universe)

    dataflows = ll.load_dataflows(dataflow_folder)
    entity_index = ll.build_entity_index(dataflows)
    name_index = ll.build_name_index(dataflows)

    used_tables = _safe_column_set(model.relationships, "FromTableName") | _safe_column_set(model.relationships, "ToTableName")
    used_tables |= _safe_column_set(model.dax_measures, "TableName")

    return {
        "model": model,
        "pbix_path": pbix_path,
        "entries": entries,
        "pbix_universe": pbix_universe,
        "direct": direct,
        "enumerators": enumerators,
        "unrecognized_dataflow_patterns": unrecognized,
        "entity_of": entity_of,
        "dataflows": dataflows,
        "entity_index": entity_index,
        "name_index": name_index,
        "guid_cache": guid_cache,
        "used_tables": used_tables,
    }


def _report_name_from_pbix(pbix_path):
    """Human-readable report name derived from the PBIX file's own name (no extension)."""
    return os.path.splitext(os.path.basename(pbix_path))[0] if pbix_path else "Report"


def _sheet_title_from_pbix(pbix_path):
    """Excel sheet titles can't contain \\/*?:[] and are capped at 31 chars."""
    name = _report_name_from_pbix(pbix_path)
    safe = re.sub(r'[\\/*?:\[\]]', " ", name).strip()
    return safe[:31] or "Report"


def format_final_source(phys):
    """Build a clean, consistent 'Final Source' text for a resolved physical
    result dict (as returned by resolve_physical_source / extract_physical_details)."""
    connector = phys.get("connector")
    if connector == "Oracle Database":
        lines = []
        if phys.get("datamart"):
            lines.append(f"Datamart = {phys['datamart']}")
        if phys.get("schema"):
            lines.append(f'Schema = "{phys["schema"]}"')
        else:
            lines.append("Schema = ??? (MISSING - requires manual override)")
        if phys.get("table"):
            lines.append(f'Item = "{phys["table"]}"')
        return "\n".join(lines) if lines else None, None, None
    if connector in ("SharePoint Excel/CSV", "Excel Workbook"):
        lines = []
        if phys.get("file"):
            lines.append(f"XLSX Name = {phys['file']}")
        if phys.get("site"):
            lines.append(f"SP = {phys['site']}")
        final = "\n".join(lines) if lines else None
        return final, phys.get("folder"), phys.get("file")
    if connector in ("Csv Document", "Web Contents"):
        lines = []
        if phys.get("file"):
            lines.append(f"File Name = {phys['file']}")
        final = "\n".join(lines) if lines else None
        return final, phys.get("folder"), phys.get("file")
    return None, None, None


def summarize_union(union_result, ctx, visited_tables_stack):
    """Return (final_source_blank, remarks_text) for a Table.Combine union row."""
    raw = union_result.get("raw", "")
    m = ll.RE_TABLE_COMBINE.search(raw)
    remarks = m.group(0) if m else raw[:300]
    return remarks


def _collect_dataflow_stems(lvl1, phys, dataflows):
    """Return the set of dataflow file stems actually reached while resolving
    one table's lineage (level1 direct binding + every hop in the chain),
    restricted to stems that correspond to a real provided dataflow file.
    This is the basis for computing which of the 102 dataflow files are
    genuinely used by this PBIX report vs. never touched."""
    stems = set()
    if lvl1 and lvl1.get("dataflow") in dataflows:
        stems.add(lvl1["dataflow"])
    for h in (phys or {}).get("hops", []):
        stem = h.get("stem") or h.get("to")
        if stem in dataflows:
            stems.add(stem)
    if phys and phys.get("resolved_stem") in dataflows:
        stems.add(phys["resolved_stem"])
    return stems


def _apply_duplicate_notices(row, ctx):
    """If a row's lineage touched a dataflow whose duplicate exports had
    genuinely differing content (collapsed to the latest in load_dataflows),
    flag it for manual confirmation instead of leaving it silently resolved."""
    notices = [ctx["dataflows"][s]["duplicate_notice"] for s in row.get("dataflow_stems", set())
               if ctx["dataflows"].get(s, {}).get("duplicate_notice")]
    if not notices:
        return
    parts = [
        f"[NEEDS MANUAL OVERRIDE - DUPLICATE DATAFLOW DETECTED] Multiple exports of "
        f"'{n['base_name']}' were found with differing content ({', '.join(n['files'])}) - "
        f"automatically selected the latest ('{n['chosen_file']}', modified {n['chosen_modified']}). "
        f"Please verify manually."
        for n in notices
    ]
    row["needs_override"] = True
    row["override_tag"] = "; ".join(filter(None, [row.get("override_tag"), "DUPLICATE DATAFLOW DETECTED"]))
    row["remarks"] = "\n".join(filter(None, [row.get("remarks")] + parts))


def resolve_table_row(table, ctx):
    """Compute the full lineage info for one PBIX table. Returns a dict of
    fields ready for the Excel row, plus a 'status' of 'found'/'unresolved'/'union'/'no_query'."""
    row = _resolve_table_row_inner(table, ctx)
    _apply_duplicate_notices(row, ctx)
    # 'unresolved'/'union' rows have no Final Source at all - a harder problem
    # than a 'found' row that's merely flagged for a quick confirmation.
    row["hard_unresolved"] = row["status"] in ("unresolved", "union")
    return row


def _resolve_table_row_inner(table, ctx):
    entries = ctx["entries"]
    expr = entries.get(table)

    if expr is None:
        return {
            "status": "no_query",
            "entities_used": f"Entity Name: {table}",
            "level1": None, "level2": None, "final_source": None,
            "folder": None, "file": None,
            "remarks": "No M/Power Query source (measures-only or calculated table).",
            "dataflow_stems": set(),
            "needs_override": False, "override_tag": None,
            "phys": None, "lvl1_raw": None, "entity": None,
        }

    members = ll.extract_table_combine_members(expr)
    union_note = None
    if members:
        # Genuine multi-source union only if EVERY combined member is itself a
        # separate top-level PBIX table/query (e.g. 30x_Dim_Factures unions 5
        # independent fact tables - no single source exists, blank by design).
        # If at least one member is a local (inline `let`-step) variable
        # instead, this is really "one dominant physical source, unioned with
        # a small local patch/exception table" (e.g. 30217-1_DEMAT_MODEL
        # unions its real Oracle source with a hand-maintained "NoDemat"
        # exception list) - fall through and resolve the table's own lineage
        # normally, noting the extra member(s) instead of blanking Final Source.
        if all(m in entries for m in members):
            m = ll.RE_TABLE_COMBINE.search(expr)
            raw_snippet = m.group(0) if m else expr[:300]
            remarks = ("[NEEDS MANUAL OVERRIDE - MULTI-SOURCE UNION] combines "
                       f"{len(members)} independent top-level sources with no single primary "
                       f"source - manual review required to confirm intended reporting source: "
                       f"{', '.join(members)}\n{raw_snippet}")
            return {
                "status": "union",
                "entities_used": f"Entity Name: {table}\n" + "\n".join(f"({m_})ref" for m_ in members),
                "level1": None, "level2": None, "final_source": None,
                "folder": None, "file": None,
                "remarks": remarks,
                "dataflow_stems": set(),
                "needs_override": True, "override_tag": "MULTI-SOURCE UNION",
                "phys": None, "lvl1_raw": None, "entity": None,
                "union_members_raw": members,
            }
        other_members = [m for m in members if m in entries]
        union_note = ("[Table.Combine union - dominant source shown; also unions local/patch "
                      f"table(s): {', '.join(other_members)}]") if other_members else \
                     "[Table.Combine union with inline-only members - dominant source shown]"

    cache = {}
    lvl1 = ll.resolve_pbix_lineage(table, ctx["pbix_universe"], ctx["direct"], ctx["entity_of"], cache, set())
    if lvl1 is None:
        unrecognized_names = {u["query"] for u in ctx.get("unrecognized_dataflow_patterns", [])}
        if ll.chain_hits_unrecognized(table, ctx["pbix_universe"], unrecognized_names):
            tag = "UNRECOGNIZED CONNECTOR SYNTAX"
            reason_text = ("A dataflow-connector call was found in this table's local M dependency "
                           "chain, but its workspace/dataflow field syntax isn't recognized by this "
                           "tool's regex yet - see the 'unrecognized dataflow pattern(s)' console "
                           "warning printed for this run.")
        else:
            tag = ll.classify_unresolved_reason("No dataflow-connector ancestor found")
            reason_text = "No dataflow-connector ancestor found in local M dependency chain."
        return {
            "status": "unresolved",
            "entities_used": f"Entity Name: {table}",
            "level1": None, "level2": None, "final_source": None,
            "folder": None, "file": None,
            "remarks": "\n".join(filter(None, [union_note,
                f"[NEEDS MANUAL OVERRIDE - {tag}] {reason_text}"])),
            "dataflow_stems": set(),
            "needs_override": True, "override_tag": tag,
            "phys": None, "lvl1_raw": None, "entity": None,
        }

    path = lvl1.get("path", [])
    entity = lvl1.get("entity") or (path[-1] if path else table)
    entities_used = f"Entity Name: {table}"
    if path:
        entities_used += "\n" + "\n".join(f"({p})ref" for p in path)

    level1_text = f"WKS Name: {lvl1['workspace']},\nDataflow: {lvl1['dataflow']}"

    phys = ll.resolve_physical_source(lvl1["dataflow"], entity, ctx["dataflows"], ctx["entity_index"],
                                       name_index=ctx["name_index"], guid_cache=ctx["guid_cache"])
    dataflow_stems = _collect_dataflow_stems(lvl1, phys, ctx["dataflows"])

    if phys.get("unresolved"):
        tag = ll.classify_unresolved_reason(phys.get("reason"))
        remarks = f"[NEEDS MANUAL OVERRIDE - {tag}] {phys.get('reason', 'Unresolved.')}"
        return {
            "status": "unresolved",
            "entities_used": entities_used,
            "level1": level1_text, "level2": None, "final_source": None,
            "folder": None, "file": None,
            "remarks": "\n".join(filter(None, [union_note, remarks])),
            "dataflow_stems": dataflow_stems,
            "needs_override": True, "override_tag": tag,
            "phys": phys, "lvl1_raw": lvl1, "entity": entity,
        }

    if phys.get("union"):
        members_desc = phys.get("raw", "")[:300]
        remarks = ("[NEEDS MANUAL OVERRIDE - MULTI-SOURCE UNION (DATAFLOW LEVEL)] entity "
                   f"'{entity}' combines {len(phys.get('members', []))} sources within the "
                   f"dataflow's own M code - manual review required: {members_desc}")
        return {
            "status": "union",
            "entities_used": entities_used,
            "level1": level1_text, "level2": None, "final_source": None,
            "folder": None, "file": None,
            "remarks": remarks,
            "dataflow_stems": dataflow_stems,
            "needs_override": True, "override_tag": "MULTI-SOURCE UNION (DATAFLOW LEVEL)",
            "phys": phys, "lvl1_raw": lvl1, "entity": entity,
        }

    hops = phys.get("hops", [])
    level2_text = None
    ambiguous_note = ""
    for h in hops:
        if h.get("level") == "level2":
            level2_text = f"WKS Name: {lvl1['workspace']},\nDataflow: {h['stem']}\n Entity: {h['entity']}"
            if h.get("ambiguous"):
                ambiguous_note = f" [Note: entity name matched {len(h['all_candidates'])} candidate files; picked '{h['stem']}' by name similarity.]"
        elif h.get("level") == "jump":
            level2_text = f"WKS Name: {h.get('workspace') or lvl1['workspace']},\nDataflow: {h['to']}\n Entity: {h['entity']}"

    final_source, folder, file_name = format_final_source(phys)

    schema_note = None
    if phys.get("connector") == "Oracle Database" and not phys.get("schema"):
        schema_note = ("[NEEDS MANUAL OVERRIDE - MISSING SCHEMA] Schema not specified anywhere in the "
                       "source dataflow's M-code for this table - cannot be auto-extracted; please "
                       "confirm the correct schema manually.")

    override_tags = []
    if union_note:
        union_note = "[NEEDS MANUAL OVERRIDE - UNION WITH LOCAL PATCH TABLE] " + union_note.lstrip("[").rstrip("]")
        override_tags.append("UNION WITH LOCAL PATCH TABLE")
    if ambiguous_note.strip():
        ambiguous_note = ("[NEEDS MANUAL OVERRIDE - AMBIGUOUS ENTITY MATCH] " +
                          ambiguous_note.strip().lstrip("[").rstrip("]") + " Please confirm correct source.")
        override_tags.append("AMBIGUOUS ENTITY MATCH")
    if schema_note:
        override_tags.append("MISSING SCHEMA")

    remarks_parts = [p for p in [union_note, ambiguous_note.strip() or None, schema_note] if p]

    return {
        "status": "found",
        "entities_used": entities_used,
        "level1": level1_text,
        "level2": level2_text,
        "final_source": final_source,
        "folder": folder,
        "file": file_name,
        "remarks": "\n".join(remarks_parts) if remarks_parts else None,
        "dataflow_stems": dataflow_stems,
        "needs_override": bool(override_tags),
        "override_tag": "; ".join(override_tags) if override_tags else None,
        "phys": phys, "lvl1_raw": lvl1, "entity": entity,
    }


def build_report(pbix_path=None, dataflow_folder=None):
    ctx = load_everything(pbix_path, dataflow_folder)
    model = ctx["model"]
    tables = list(model.tables)

    rows = []
    for t in sorted(tables):
        info = resolve_table_row(t, ctx)
        is_used = t in ctx["used_tables"]
        info["table"] = t
        info["is_used"] = is_used
        rows.append(info)

    return rows, ctx


def write_workbook(rows, ctx, output_path=None):
    output_path = output_path or OUTPUT_PATH

    # Ensure output directory exists before attempting to write
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.isdir(output_dir):
        try:
            os.makedirs(output_dir, exist_ok=True)
        except Exception as e:
            raise IOError(f"Cannot create output directory '{output_dir}': {e}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _sheet_title_from_pbix(ctx.get("pbix_path"))

    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)

    ws.cell(row=2, column=1, value="Found Source")
    ws.cell(row=2, column=1).fill = GREEN_FILL
    ws.cell(row=3, column=1, value="Unresolved / No Source Found")
    ws.cell(row=3, column=1).fill = GREY_FILL
    ws.cell(row=4, column=1, value="Tables not used")
    ws.cell(row=4, column=1).fill = RED_FILL
    ws.cell(row=5, column=1, value="Needs Manual Override (source found, please confirm)")
    ws.cell(row=5, column=1).fill = YELLOW_FILL
    ws.cell(row=2, column=2, value=f"Report: {_report_name_from_pbix(ctx.get('pbix_path'))}\nDownloaded: WKS DTF FINANCE")

    r = 6
    dataflows_used_l1 = set()
    dataflows_used_l2 = set()
    unused_tables = []

    for info in rows:
        for c in range(1, 10):
            ws.cell(row=r, column=c).alignment = WRAP

        ws.cell(row=r, column=3, value=info["entities_used"])
        ws.cell(row=r, column=4, value=info["level1"])
        ws.cell(row=r, column=5, value=info["level2"])
        ws.cell(row=r, column=6, value=info["final_source"])
        ws.cell(row=r, column=7, value=info["folder"])
        ws.cell(row=r, column=8, value=info["file"])
        ws.cell(row=r, column=9, value=info["remarks"])

        if not info["is_used"]:
            fill = RED_FILL
            unused_tables.append(info["table"])
        elif info.get("hard_unresolved"):
            fill = GREY_FILL
        elif info.get("needs_override"):
            fill = YELLOW_FILL
        elif info["status"] == "found":
            fill = GREEN_FILL
        elif info["status"] == "no_query":
            fill = None
        else:
            fill = GREY_FILL

        if fill:
            for c in range(3, 7):
                ws.cell(row=r, column=c).fill = fill

        if info["level1"]:
            dataflows_used_l1.add(info["level1"].split("Dataflow: ", 1)[-1])
        if info["level2"]:
            dataflows_used_l2.add(info["level2"].split("Dataflow: ", 1)[-1].split("\n")[0])

        r += 1

    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 35
    ws.column_dimensions["H"].width = 25
    ws.column_dimensions["I"].width = 45

    ws2 = wb.create_sheet("Dataflows used")
    ws2.cell(row=1, column=1, value="Level 1 - Dataflows").font = Font(bold=True)
    for i, d in enumerate(sorted(dataflows_used_l1), start=2):
        ws2.cell(row=i, column=1, value=f"#{i - 1} {d}")
    start2 = len(dataflows_used_l1) + 3
    ws2.cell(row=start2, column=1, value="Level 2 - Dataflows").font = Font(bold=True)
    for i, d in enumerate(sorted(dataflows_used_l2), start=start2 + 1):
        ws2.cell(row=i, column=1, value=d)
    ws2.column_dimensions["A"].width = 60

    ws3 = wb.create_sheet("Unused tables")
    for i, t in enumerate(unused_tables, start=1):
        ws3.cell(row=i, column=1, value=t)
    ws3.column_dimensions["A"].width = 45

    # --- Dataflow File Coverage sheet -------------------------------------
    # This is the accurate, hop-aware answer to "which of the N provided
    # dataflow JSON files are actually reached by this PBIX report's
    # lineage?" (as opposed to the Level1/Level2 sheet above, which is a
    # best-effort text parse of only the first two hops). It walks every
    # dataflow_stems set collected while resolving each table (level1 direct
    # binding + every hop: level2/jump/level2-guid), so multi-hop chains are
    # fully credited, then diffs against every dataflow file actually present
    # on disk to flag files with zero reachability from any PBIX table.
    all_used_stems = set()
    for info in rows:
        all_used_stems |= info.get("dataflow_stems", set())
    all_stems = set(ctx["dataflows"].keys())
    unused_stems = all_stems - all_used_stems

    ws4 = wb.create_sheet("Dataflow File Coverage")
    ws4.cell(row=1, column=1, value="Dataflow File").font = Font(bold=True)
    ws4.cell(row=1, column=2, value="Reached by PBIX lineage?").font = Font(bold=True)
    r4 = 2
    for stem in sorted(all_stems):
        used = stem in all_used_stems
        ws4.cell(row=r4, column=1, value=stem)
        ws4.cell(row=r4, column=2, value="Used" if used else "Unused")
        if not used:
            ws4.cell(row=r4, column=1).fill = RED_FILL
            ws4.cell(row=r4, column=2).fill = RED_FILL
        r4 += 1
    ws4.column_dimensions["A"].width = 60
    ws4.column_dimensions["B"].width = 22

    btr.add_transformations_sheet(wb, rows, ctx)

    try:
        wb.save(output_path)
    except Exception as e:
        raise IOError(f"Cannot write to '{output_path}': {e}")
    print(f"Saved: {output_path}")

    unrecognized = ctx.get("unrecognized_dataflow_patterns") or []
    if unrecognized:
        print(f"WARNING: {len(unrecognized)} quer(y/ies) call a dataflow connector "
              "(PowerPlatform.Dataflows/Dataflows.Contents/PowerBI.Dataflows) using an M syntax "
              "this tool doesn't recognize - affected rows are tagged 'UNRECOGNIZED CONNECTOR "
              "SYNTAX' instead of being resolved. lineage_lib.py's extract_fields() (RE_FIELD_EQ_A / "
              "RE_RECORD_SELECTOR) needs a new style added for:")
        for u in unrecognized:
            print(f"  - {u['query']}: {u['snippet'][:120]}")

    print(f"Total rows: {len(rows)}")
    found = sum(1 for i in rows if i["status"] == "found")
    unresolved = sum(1 for i in rows if i["status"] == "unresolved")
    union = sum(1 for i in rows if i["status"] == "union")
    no_query = sum(1 for i in rows if i["status"] == "no_query")
    print(f"  found={found} unresolved={unresolved} union={union} no_query={no_query}")
    print(f"  unused (not in relationships/measures)={len(unused_tables)}")
    print(f"  dataflow files provided={len(all_stems)}, reached by PBIX lineage={len(all_used_stems)}, "
          f"never touched={len(unused_stems)}")

    needs_override = sum(1 for i in rows if i.get("needs_override") and not i.get("hard_unresolved"))
    hard_unresolved = sum(1 for i in rows if i.get("hard_unresolved"))
    print(f"  needs_manual_override (soft - source found, please confirm)={needs_override} "
          f"({needs_override / len(rows):.1%} of total)")
    print(f"  hard_unresolved (no source found at all)={hard_unresolved} ({hard_unresolved / len(rows):.1%} of total)")
    tag_counts = Counter(i["override_tag"] for i in rows if i.get("needs_override") and not i.get("hard_unresolved"))
    for tag, cnt in sorted(tag_counts.items(), key=lambda x: -x[1]):
        print(f"    - {tag}: {cnt}")


if __name__ == "__main__":
    rows, ctx = build_report()
    write_workbook(rows, ctx)

    import build_dataflow_table_lineage_report as dtlr
    dtlr.build_and_save(ctx)
