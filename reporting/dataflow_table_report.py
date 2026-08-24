"""
build_dataflow_table_lineage_report.py

Generates Dataflow_Table_Lineage_Report.xlsx - a companion report to
Generated_CashPlus_Lineage.xlsx, in the "Overview" / "Table Lineage" column
layout (same exact column names as the original hand-built report), with
every value derived from the PBIX file + dataflow JSON files configured in
config.py (via the shared lineage_lib.py engine - no separate/duplicated
resolution logic).

There is no target/reference workbook involved: this report is meant to be
authoritative on its own. Anything the engine cannot resolve with full
confidence (dataflow not found, ambiguous match, missing schema, multi-source
union, etc.) is flagged in the "Status" column as NEEDS MANUAL REVIEW, with
the specific issue type named in "Unresolved Reason" and the row highlighted
yellow - instead of being silently left blank or compared against a target.

Usage:
    python build_dataflow_table_lineage_report.py
"""
import sys
import os

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import Counter

from core import lineage_lib as ll
from reporting import lineage_report as blr
import config

OUTPUT_PATH = config.DATAFLOW_LINEAGE_XLSX

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")

HEADERS = [
    "Table Name (Report)", "Is Calculated / No External Source", "Status",
    "Primary Source System", "Schema / SharePoint Site", "Table Name / File Name",
    "Sheet / Tab Name", "Union Members (if multi-source)", "Secondary / Lookup Source(s)",
    "Resolution Method", "Unresolved Reason", "Lineage Path",
]


def _short_source_desc(phys):
    """One-line 'connector (key details)' description of a resolved physical
    source dict, used to describe individual union members."""
    if not phys:
        return "unknown"
    if phys.get("unresolved"):
        return f"unresolved ({(phys.get('reason') or '')[:100]})"
    if phys.get("union"):
        return f"Union of {len(phys.get('members', []))} sources"
    c = phys.get("connector")
    if c == "Oracle Database":
        return f"Oracle Database (Schema={phys.get('schema')}, Table={phys.get('table')})"
    if c in ("SharePoint Excel/CSV", "Excel Workbook"):
        return f"{c} (File={phys.get('file')})"
    if c in ("Csv Document", "Web Contents"):
        return f"{c} (File={phys.get('file')})"
    return c or "unknown connector"


def resolve_member_source(member, ctx):
    """Resolve a Table.Combine union member as its own top-level PBIX table.
    Members that are local (non-top-level) variables aren't separately
    resolvable - reported as such rather than guessed."""
    entries = ctx["entries"]
    if member not in entries:
        return {"unresolved": True,
                "reason": f"'{member}' is a local (non-top-level) variable, not a separate PBIX table/query."}
    lvl1 = ll.resolve_pbix_lineage(member, ctx["pbix_universe"], ctx["direct"], ctx["entity_of"], {}, set())
    if lvl1 is None:
        return {"unresolved": True, "reason": "No dataflow-connector ancestor found."}
    path = lvl1.get("path", [])
    entity = lvl1.get("entity") or (path[-1] if path else member)
    return ll.resolve_physical_source(lvl1["dataflow"], entity, ctx["dataflows"], ctx["entity_index"],
                                       name_index=ctx["name_index"], guid_cache=ctx["guid_cache"])


def build_row(table, ctx):
    info = blr.resolve_table_row(table, ctx)
    status = info["status"]

    row = {
        "table": table, "is_calc": "No", "status_label": None,
        "primary_source": None, "schema_site": None, "table_file": None, "sheet": None,
        "union_members": None, "secondary": None, "resolution_method": None,
        "unresolved_reason": None, "lineage_path": None,
        "needs_override": info.get("needs_override", False),
        "override_tag": info.get("override_tag"),
    }

    if status == "no_query":
        row["is_calc"] = "Yes"
        row["status_label"] = "Calculated"
        row["lineage_path"] = f"[PBIX] '{table}' -> calculated/static table (no external source, no Power Query expression)."
        return row

    if status == "union" and info.get("union_members_raw") is not None:
        members = info["union_members_raw"]
        descs = [f"{m} -> {_short_source_desc(resolve_member_source(m, ctx))}" for m in members]
        row["status_label"] = "NEEDS MANUAL REVIEW"
        row["primary_source"] = "Union of multiple sources"
        row["union_members"] = "; ".join(descs)
        row["resolution_method"] = "Table.Combine union of independent top-level queries (no single dominant physical source)"
        row["unresolved_reason"] = (f"[{row['override_tag']}] combines {len(members)} independent top-level sources "
                                     f"with no single primary source - manual review required to confirm intended "
                                     f"reporting source: {', '.join(members)}")
        row["lineage_path"] = (f"[PBIX] '{table}' -> Table.Combine union of top-level queries: {members} "
                                f"(multi-source; no single physical source - flagged for manual review).")
        return row

    if status == "union" and info.get("phys") is not None:
        lvl1, entity, phys = info["lvl1_raw"], info["entity"], info["phys"]
        members = phys.get("members", [])
        row["status_label"] = "NEEDS MANUAL REVIEW"
        row["primary_source"] = "Union of multiple sources"
        row["union_members"] = "; ".join(f"{name} -> {_short_source_desc(sub)}" for name, sub in members)
        row["resolution_method"] = f"{lvl1.get('method', '')}; Table.Combine union within dataflow-level query '{entity}'"
        row["unresolved_reason"] = (f"[{row['override_tag']}] entity '{entity}' in dataflow '{lvl1['dataflow']}' "
                                     f"combines {len(members)} sources within the dataflow's own M code - manual review required.")
        row["lineage_path"] = (f"[PBIX] '{table}' -> dataflow '{lvl1['dataflow']}' (workspace: {lvl1['workspace']}) "
                                f":: entity '{entity}' -> Table.Combine UNION of {len(members)} sources: "
                                f"{[name for name, _ in members]}")
        return row

    if status == "unresolved":
        row["status_label"] = "NEEDS MANUAL REVIEW"
        lvl1 = info.get("lvl1_raw")
        if lvl1:
            reason = (info.get("phys") or {}).get("reason", "Unresolved.")
            row["resolution_method"] = lvl1.get("method")
            row["lineage_path"] = (f"[PBIX] '{table}' -> dataflow '{lvl1['dataflow']}' (workspace: {lvl1['workspace']}) "
                                    f":: entity '{info.get('entity')}' -> UNRESOLVED: {reason}")
        else:
            reason = "No dataflow-connector ancestor found in local M dependency chain."
            row["lineage_path"] = f"[PBIX] '{table}' -> no dataflow-connector ancestor found in local M dependency chain."
        row["unresolved_reason"] = f"[{row['override_tag']}] {reason}"
        return row

    # status == "found"
    lvl1, entity, phys = info["lvl1_raw"], info["entity"], info["phys"]
    connector = phys.get("connector")
    row["status_label"] = "NEEDS MANUAL REVIEW" if row["needs_override"] else "Resolved"
    row["primary_source"] = connector
    row["resolution_method"] = lvl1.get("method")

    if connector == "Oracle Database":
        row["schema_site"] = phys.get("schema")
        row["table_file"] = phys.get("table")
    elif connector in ("SharePoint Excel/CSV", "Excel Workbook"):
        row["schema_site"] = phys.get("site")
        row["table_file"] = phys.get("file")
    elif connector in ("Csv Document", "Web Contents"):
        row["table_file"] = phys.get("file")

    if row["needs_override"]:
        remark = (info.get("remarks") or "").replace("\n", " ")
        row["unresolved_reason"] = f"[{row['override_tag']}] {remark}"

    parts = [f"[PBIX] '{table}' -> dataflow '{lvl1['dataflow']}' (workspace: {lvl1['workspace']}) :: entity '{entity}'"]
    for h in phys.get("hops", []):
        lvl = h.get("level")
        if lvl == "level2":
            note = " [ambiguous match]" if h.get("ambiguous") else ""
            parts.append(f"=> linked entity resolved in dataflow '{h['stem']}'{note}")
        elif lvl == "level2-guid":
            parts.append(f"=> linked entity resolved via GUID lookup to dataflow '{h['stem']}' (dataflow name: {h.get('dataflow_name')})")
        elif lvl == "jump":
            parts.append(f"=> in-document jump to dataflow '{h['to']}', entity '{h['entity']}'")
    parts.append(f"-> {connector} [{_short_source_desc(phys)}]")
    row["lineage_path"] = "  ".join(parts)
    return row


def build_rows(ctx):
    model = ctx["model"]
    return [build_row(t, ctx) for t in sorted(model.tables)]


def write_overview_sheet(wb, rows):
    ws = wb.create_sheet("Overview", 0)
    ws.cell(row=2, column=2, value="PBIX Table -> Dataflow -> Physical Source Lineage").font = Font(bold=True, size=13)
    ws.cell(row=3, column=2, value="Traced from the PBIX report model through the dataflow chain "
                                    "(linked entities followed) to the physical connector.")
    r = 5
    ws.cell(row=r, column=2, value="Total tables in report model:")
    ws.cell(row=r, column=3, value=len(rows))
    r += 1
    status_counts = Counter(row["status_label"] for row in rows)
    for label in ("Resolved", "Calculated", "NEEDS MANUAL REVIEW"):
        ws.cell(row=r, column=2, value=f"{label}:")
        ws.cell(row=r, column=3, value=status_counts.get(label, 0))
        if label == "NEEDS MANUAL REVIEW":
            ws.cell(row=r, column=2).fill = YELLOW_FILL
            ws.cell(row=r, column=3).fill = YELLOW_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="Source systems identified:").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=2, value="Source System").font = Font(bold=True)
    ws.cell(row=r, column=3, value="Table Count").font = Font(bold=True)
    r += 1
    src_counts = Counter(row["primary_source"] for row in rows if row["primary_source"])
    for src, cnt in src_counts.most_common():
        ws.cell(row=r, column=2, value=src)
        ws.cell(row=r, column=3, value=cnt)
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="NEEDS MANUAL REVIEW - by issue type:").font = Font(bold=True)
    r += 1
    tag_counts = Counter(row["override_tag"] for row in rows if row.get("needs_override"))
    for tag, cnt in tag_counts.most_common():
        ws.cell(row=r, column=2, value=tag)
        ws.cell(row=r, column=3, value=cnt)
        r += 1

    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15


def write_table_lineage_sheet(wb, rows):
    ws = wb.create_sheet("Table Lineage")
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFFFF")

    r = 2
    for row in rows:
        vals = [row["table"], row["is_calc"], row["status_label"], row["primary_source"],
                row["schema_site"], row["table_file"], row["sheet"], row["union_members"],
                row["secondary"], row["resolution_method"], row["unresolved_reason"], row["lineage_path"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP
            if row.get("needs_override"):
                cell.fill = YELLOW_FILL
        r += 1

    widths = [30, 16, 20, 20, 18, 26, 16, 45, 35, 35, 55, 80]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_workbook(rows, output_path=None):
    output_path = output_path or OUTPUT_PATH
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_overview_sheet(wb, rows)
    write_table_lineage_sheet(wb, rows)
    wb.save(output_path)
    print(f"Saved: {output_path}")
    status_counts = Counter(row["status_label"] for row in rows)
    print(f"Total rows: {len(rows)}  " + "  ".join(f"{k}={v}" for k, v in status_counts.items()))
    tag_counts = Counter(row["override_tag"] for row in rows if row.get("needs_override"))
    for tag, cnt in sorted(tag_counts.items(), key=lambda x: -x[1]):
        print(f"    - {tag}: {cnt}")


def build_and_save(ctx=None, pbix_path=None, dataflow_folder=None, output_path=None):
    if ctx is None:
        ctx = blr.load_everything(pbix_path, dataflow_folder)
    rows = build_rows(ctx)
    write_workbook(rows, output_path)
    return rows


if __name__ == "__main__":
    build_and_save()
