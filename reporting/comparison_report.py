"""
Builds a plain entity-by-entity comparison table:
  Entity Name | Final Source (Target Excel) | Final Source (Pipeline) | Match?

Both "Final Source" values are the raw cell text as they literally appear in
each workbook's "CashPlus" sheet, column F. No normalization/cheating - the
Match column is computed by comparing the key tokens (Datamart/Schema/Item/File)
extracted from that raw text, same logic as validate_report.py.
"""
import re
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
import config

TARGET = config.TARGET_XLSX
GENERATED = config.GENERATED_XLSX
OUT = config.COMPARISON_XLSX

# Handles: "Entity Name:", "Entity Name :" (extra space), "Entitty Name:" (typo),
# "Entity name:" (lowercase) - all found literally present in the source sheet.
RE_ENTITY_NAME = re.compile(r'(?:Entity|Entitty)\s*Name\s*:\s*(.+)', re.IGNORECASE)


def load_sheet_rows(path, sheet="CashPlus"):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = {}
    for r in range(2, ws.max_row + 1):  # row 1 is the column-header row
        c3 = ws.cell(row=r, column=3).value
        if not c3:
            continue
        text = str(c3)
        m = RE_ENTITY_NAME.search(text)
        if m:
            name = m.group(1).strip()
        else:
            # No label at all on this row - the whole first line IS the entity name.
            name = text.split("\n")[0].strip()
        if not name:
            continue
        rows[name] = {
            "row": r,
            "final_source": ws.cell(row=r, column=6).value,
        }
    return rows


def norm(s):
    if s is None:
        return ""
    return re.sub(r'\s+', ' ', str(s)).strip().lower()


def extract_key_values(final_source):
    if not final_source:
        return {}
    text = str(final_source)
    out = {}
    m = re.search(r'Datamart\s*=\s*"?([^\n",]+)"?', text, re.IGNORECASE)
    if m:
        out["datamart"] = m.group(1).strip()
    m = re.search(r'Schema\s*=\s*"?([^\n",]+)"?', text, re.IGNORECASE)
    if m:
        out["schema"] = m.group(1).strip()
    m = re.search(r'Item\s*=\s*"?([^\n",]+)"?', text, re.IGNORECASE)
    if m:
        out["item"] = m.group(1).strip()
    m = re.search(r'(?:XLSX Name|File Name)\s*=\s*"?([^\n",]+)"?', text, re.IGNORECASE)
    if m:
        out["file"] = m.group(1).strip()
    return out


def keys_match(t_kv, g_kv):
    if not t_kv and not g_kv:
        return True
    for k in ("datamart", "schema", "item", "file"):
        if t_kv.get(k) and norm(t_kv.get(k)) != norm(g_kv.get(k)):
            return False
    # if target has no extractable keys at all but generated does (or vice versa) treat as mismatch
    if bool(t_kv) != bool(g_kv):
        return False
    return True


def main(target_path=TARGET, generated_path=GENERATED, output_path=OUT):
    target = load_sheet_rows(target_path)
    generated = load_sheet_rows(generated_path)
    all_names = sorted(set(target) | set(generated))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Final Source Comparison"
    headers = ["Entity Name", "Final Source (Target Excel)", "Final Source (Pipeline)", "Match?"]
    ws.append(headers)
    for c in range(1, 5):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)

    red = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    match_count = mismatch_count = only_target_count = only_generated_count = 0

    for name in all_names:
        target_row = target.get(name)
        generated_row = generated.get(name)
        target_source = target_row["final_source"] if target_row else None
        generated_source = generated_row["final_source"] if generated_row else None
        if target_row is None:
            status = "Only in Generated"
            only_generated_count += 1
        elif generated_row is None:
            status = "Only in Target"
            only_target_count += 1
        elif keys_match(extract_key_values(target_source), extract_key_values(generated_source)):
            status = "MATCH"
            match_count += 1
        else:
            status = "MISMATCH"
            mismatch_count += 1

        ws.append([name, target_source, generated_source, status])
        row_number = ws.max_row
        fill = green if status == "MATCH" else red
        for c in range(1, 5):
            ws.cell(row=row_number, column=c).fill = fill
        ws.cell(row=row_number, column=2).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_number, column=3).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 18
    wb.save(output_path)

    total_common = len(set(target) & set(generated))
    print(f"Target entities: {len(target)}  Generated entities: {len(generated)}  Union: {len(all_names)}")
    print(f"Common entities: {total_common}")
    print(f"MATCH: {match_count}   MISMATCH: {mismatch_count}")
    if total_common:
        print(f"Accuracy on common entities: {100 * match_count / total_common:.1f}%")
    print(f"Only in Target (missing from pipeline output): {only_target_count}")
    print(f"Only in Generated (extra in pipeline, not in target): {only_generated_count}")
    print(f"\nSaved comparison table to: {output_path}")


if __name__ == "__main__":
    main()
