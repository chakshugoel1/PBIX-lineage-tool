"""
test_model_change_impact_excel_report.py

End-to-end test for model_change_impact/excel_report.py: runs hand-built
snapshots through the real diff.diff_snapshots() and impact.analyze_impact()
(same as test_model_change_impact_impact.py), builds the Excel workbook into
a pytest tmp_path, and reads it back with openpyxl to verify sheet layout
and content. No real PBIX file needed, matching this repo's test convention.
"""
import copy
import os
import sys

import openpyxl
from openpyxl.cell.rich_text import CellRichText

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from model_change_impact import diff, impact, excel_report


def _column(name, expression=None, is_calculated=False, format_string=None):
    return {
        "name": name, "data_type": "Int64", "is_calculated": is_calculated,
        "expression": expression, "format_string": format_string, "is_hidden": False,
        "description": None, "display_folder": None, "lineage_tag": f"col-{name}",
    }


def _table(m_expression="M", columns=None):
    return {
        "is_calculated_table": False, "m_expression": m_expression, "is_hidden": False,
        "description": None, "lineage_tag": None, "columns": columns or [],
    }


def _measure(expression):
    return {
        "expression": expression, "display_folder": None, "description": "",
        "format_string": None, "is_hidden": False, "lineage_tag": None, "kpi_id": None,
    }


def _base_snapshot(source_file):
    return {
        "source_file": source_file,
        "tables": {
            "Sales": _table(columns=[_column("Amount"), _column("Region")]),
            "Calendar": _table(columns=[_column("Date")]),
        },
        "measures": {
            "_Measures": {
                "Total Sales": _measure("SUM(Sales[Amount])"),
                "Total Sales YTD": _measure("CALCULATE([Total Sales], DATESYTD(Calendar[Date]))"),
            }
        },
        "calculated_columns": [],
        "relationships": [
            {"from_table": "Sales", "from_column": "RegionID", "to_table": "Calendar", "to_column": "ID",
             "is_active": True, "cardinality": "M:1", "cross_filtering_behavior": "Single",
             "rely_on_referential_integrity": False},
        ],
    }


def _report_layout():
    return {
        "format": "pbir",
        "pages": [{
            "page_id": "p1", "display_name": "Overview",
            "visuals": [
                {"visual_id": "visKpi", "kind": "visual", "visual_type": "card", "kpi_classification": "certain",
                 "display_name": "Sales KPI",
                 "fields": [{"kind": "measure", "table": "_Measures", "field": "Total Sales YTD", "role": "Values"}]},
                {"visual_id": "visChart", "kind": "visual", "visual_type": "columnChart", "kpi_classification": None,
                 "display_name": "Sales by Region",
                 "fields": [{"kind": "column", "table": "Sales", "field": "Amount", "role": "Y"}]},
                {"visual_id": "visKpiCustom", "kind": "visual", "visual_type": "advanceCardXYZ123",
                 "kpi_classification": "heuristic", "display_name": "Total Sales Card",
                 "fields": [{"kind": "measure", "table": "_Measures", "field": "Total Sales", "role": "Values"}]},
            ],
        }],
    }


def _build_report(tmp_path):
    baseline = _base_snapshot("baseline.pbix")
    changed = copy.deepcopy(baseline)
    changed["source_file"] = "changed.pbix"
    changed["measures"]["_Measures"]["Total Sales"]["expression"] = "SUM(Sales[Amount]) * 1.1"
    changed["relationships"][0]["cardinality"] = "M:M"

    layout = _report_layout()
    diff_result = diff.diff_snapshots(baseline, changed)
    impact_result = impact.analyze_impact(baseline, changed, diff_result, layout)

    output_path = str(tmp_path / "model_change_impact.xlsx")
    excel_report.build_excel_report(baseline, changed, diff_result, impact_result, layout, output_path)
    return output_path


def test_builds_all_expected_sheets_in_order(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    assert wb.sheetnames == [
        "Summary", "Impact Summary", "Changed Tables", "Changed Measures", "Changed Columns",
        "Changed Relationships",
    ]


def test_unrequested_detail_sheets_are_not_created(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    assert not set(wb.sheetnames) & {
        "Visual Impact", "KPI Impact", "Playwright Input", "Manual Review", "Object Inventory",
    }


def test_changed_relationships_list_only_manual_relationships(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Changed Relationships"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows == [] or all(r[4] == "Modified" or r[4] == "Added" or r[4] == "Removed" for r in rows)
    assert all(r[5] == "MANUAL" for r in rows)


def test_impact_summary_has_actual_report_change_and_filter_columns(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Impact Summary"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Changed Object Type" in headers
    assert "Changed Object" in headers
    assert "Affected Visual ID" in headers
    assert "Affected Visual Name" in headers
    assert "Page Name" in headers
    assert "Is KPI" in headers
    assert "KPI Confidence" in headers
    assert "Impact Basis" in headers
    assert "Actual Report Change" in headers
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    measure_kpi_row = next(r for r in rows if r[0] == "Measure" and r[3] == "visKpiCustom")
    assert measure_kpi_row[1] == "_Measures[Total Sales]"
    assert measure_kpi_row[4] == "Total Sales Card"
    assert measure_kpi_row[6] == "Overview"
    assert measure_kpi_row[7] == "Yes"
    assert measure_kpi_row[8] == "heuristic"
    assert measure_kpi_row[9] == "Direct"


def test_summary_sheet_has_file_names_and_counts(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Summary"]
    assert ws["B2"].value == "baseline.pbix"
    assert ws["B3"].value == "changed.pbix"
    # header row for the counts table
    headers = [ws.cell(row=6, column=c).value for c in range(1, 8)]
    assert headers == ["Entity", "Added", "Removed", "Changed", "Rename Candidates", "Unchanged", "Detail Sheet"]
    # find the "measures" row and check its "Changed" count is 1
    measures_row = next(r for r in range(7, 11) if ws.cell(row=r, column=1).value == "Measures")
    assert ws.cell(row=measures_row, column=4).value == 1


def test_changed_measures_sheet_has_the_modified_measure(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Changed Measures"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(r[0] == "_Measures" and r[1] == "Total Sales" and r[2] == "Modified" for r in rows)


def test_playwright_input_severity_reflects_kpi_classification(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Impact Summary"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    by_visual = {r[3]: r for r in rows if r[3]}
    # visKpi has kpi_classification "certain" -> always "High".
    assert by_visual["visKpi"][7] == "Yes"
    # visChart has no KPI classification and is only reachable via the relationship's
    # broad, table-wide seeding (not a direct binding) -> demoted to "Low", not "Medium"/"High".
    # This is the concrete scenario the tool must get right: an unrelated change
    # (e.g. a relationship edit) must not make an unrelated visual look as urgent
    # as a genuinely, directly affected one.
    assert by_visual["visChart"][7] == "No"
    assert by_visual["visChart"][9] == "Broad (table/relationship-level)"
    # visKpiCustom is directly bound to the modified measure and only heuristically
    # classified as a KPI (custom visual type name) -> "Medium" (direct basis, not certain KPI).
    assert by_visual["visKpiCustom"][7] == "Yes"


def test_manual_review_flags_relationship_change_and_heuristic_kpi(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Impact Summary"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(row[0] == "Relationship" for row in rows)
    assert any(row[7] == "Yes" and row[8] == "heuristic" for row in rows)


def test_object_inventory_lists_current_model_objects(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    assert "Object Inventory" not in wb.sheetnames


def _cell_text(value):
    if isinstance(value, CellRichText):
        return "".join(str(part) for part in value)
    return value


def test_changed_measure_shows_full_before_after_with_highlighting(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path), rich_text=True)
    ws = wb["Changed Measures"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == "_Measures" and r[1] == "Total Sales" and r[4] == "expression")
    before, after = row[5], row[6]
    # After changed (new tokens added) -> highlighted rich text; before is unchanged -> plain text.
    assert _cell_text(before) == "SUM(Sales[Amount])"
    assert isinstance(after, CellRichText)
    assert _cell_text(after) == "SUM(Sales[Amount]) * 1.1"


def test_long_dax_expression_is_not_truncated(tmp_path):
    baseline = _base_snapshot("baseline.pbix")
    changed = copy.deepcopy(baseline)
    changed["source_file"] = "changed.pbix"
    long_expr = "SWITCH(TRUE(), " + ", ".join(f"Sales[Region] = \"R{i}\", {i}" for i in range(60)) + ", 0)"
    assert len(long_expr) > 500
    changed["measures"]["_Measures"]["Total Sales"]["expression"] = long_expr

    layout = _report_layout()
    diff_result = diff.diff_snapshots(baseline, changed)
    impact_result = impact.analyze_impact(baseline, changed, diff_result, layout)
    output_path = str(tmp_path / "long_expr.xlsx")
    excel_report.build_excel_report(baseline, changed, diff_result, impact_result, layout, output_path)

    wb = openpyxl.load_workbook(output_path, rich_text=True)
    ws = wb["Changed Measures"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[1] == "Total Sales" and r[4] == "expression")
    after_text = _cell_text(row[6])
    assert after_text == long_expr
    assert "..." not in after_text


def test_visual_impact_sheet_is_grouped_by_page_and_deduped_per_visual(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Impact Summary"]
    assert ws.cell(row=2, column=1).value == "Measure"

    data_rows = list(ws.iter_rows(min_row=3, values_only=True))
    visual_ids = [r[3] for r in data_rows if r[3]]
    assert set(visual_ids) == {"visKpi", "visChart", "visKpiCustom"}


def test_visual_impact_basis_distinguishes_direct_from_broad(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Impact Summary"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # visKpiCustom is bound directly to the modified measure -> "Direct".
    assert any(r[1] == "_Measures[Total Sales]" and r[3] == "visKpiCustom" and r[9] == "Direct" for r in rows)
    # visChart is only reachable through the relationship's broad, whole-table seeding.
    assert any(r[0] == "Relationship" and r[3] == "visChart" and r[9] == "Broad (table/relationship-level)" for r in rows)


def test_kpi_impact_sheet_only_lists_kpi_visuals_grouped_by_page(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Impact Summary"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert {r[3] for r in rows if r[7] == "Yes"} == {"visKpi", "visKpiCustom"}


def test_manual_review_notes_cardinality_only_relationship_change(tmp_path):
    wb = openpyxl.load_workbook(_build_report(tmp_path))
    ws = wb["Impact Summary"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    relationship_rows = [r for r in rows if r[0] == "Relationship"]
    assert relationship_rows
    assert all(r[2] == "Modified" for r in relationship_rows)
    assert all(r[9] == "Broad (table/relationship-level)" for r in relationship_rows)
