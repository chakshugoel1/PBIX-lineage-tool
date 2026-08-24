import openpyxl

from reporting import lineage_report as report


def _row(**overrides):
    row = {
        "status": "found",
        "is_used": True,
        "hard_unresolved": False,
        "needs_override": False,
        "entities_used": "Entity Name: Sales",
        "level1": "Level 1",
        "level2": "Level 2",
        "final_source": "Final Source",
        "folder": "Folder",
        "file": "File",
        "remarks": "Remark",
        "dataflow_stems": set(),
    }
    row.update(overrides)
    return row


def test_main_report_uses_requested_column_order(monkeypatch, tmp_path):
    monkeypatch.setattr(report.btr, "add_transformations_sheet", lambda workbook, rows, ctx: None)
    output_path = tmp_path / "main.xlsx"
    report.write_workbook([_row()], {"pbix_path": "CashPlus.pbix", "dataflows": {}}, str(output_path))

    worksheet = openpyxl.load_workbook(output_path).active
    headers = [worksheet.cell(row=1, column=column).value for column in range(1, 10)]
    assert headers == [
        "Report Name", "Status", "Remarks", "Entities Used across report",
        "Source - Level 1", "Source - Level 2", "Final Source", "Folder PATH", "File Name",
    ]
    assert worksheet.cell(row=2, column=1).value == "Report: CashPlus\nDownloaded: WKS DTF FINANCE"
    assert worksheet.cell(row=2, column=2).value == "Found Source [Automatically]"
    assert worksheet.cell(row=2, column=3).value == "Remark"
    assert worksheet.cell(row=2, column=4).value == "Entity Name: Sales"
    assert worksheet.cell(row=2, column=5).value == "Level 1"


def test_unrecognized_connector_is_described_as_m_query_review(monkeypatch):
    context = {
        "entries": {"Sales": "let Source = PowerPlatform.Dataflows(null) in Source"},
        "pbix_universe": object(),
        "direct": {},
        "entity_of": {},
        "unrecognized_dataflow_patterns": [{"query": "Sales", "snippet": "..."}],
    }
    monkeypatch.setattr(report.ll, "resolve_pbix_lineage", lambda *args: None)
    monkeypatch.setattr(report.ll, "chain_hits_unrecognized", lambda *args: True)

    result = report.resolve_table_row("Sales", context)

    assert result["override_tag"] == "M QUERY REFERENCE NOT EXPLICIT"
    assert "dynamically constructed, indirect, or not explicitly stated" in result["remarks"]
    assert "tool" not in result["remarks"].lower()


def test_report_name_is_repeated_for_each_data_row(monkeypatch, tmp_path):
    monkeypatch.setattr(report.btr, "add_transformations_sheet", lambda workbook, rows, ctx: None)
    output_path = tmp_path / "main.xlsx"
    report.write_workbook([_row(), _row(entities_used="Entity Name: Finance")],
                          {"pbix_path": "CashPlus.pbix", "dataflows": {}}, str(output_path))

    worksheet = openpyxl.load_workbook(output_path).active
    assert worksheet.cell(row=2, column=1).value == worksheet.cell(row=3, column=1).value
